"""
Management command для импорта прайс-листа поставщика ("прайс для Родиона.xlsx").

Прайс — это таблица из двух колонок: название товара и цена.
Строки без цены — СЕКЦИИ (заголовки групп товаров), например "HDD", "Monitors TFT".

Правила обработки:
1. Секция → категория (существующая или создаётся новая).
2. Бренд определяется из названия по словарю (существующие + новые из прайса).
3. Характеристики извлекаются из названия (EAV): диагональ, разрешение, частота,
   тип матрицы, CPU, RAM, SSD, VRAM, сокет, тип ОЗУ, мощность БП, булевы признаки.
4. Теги: gaming, wireless, rgb, mechanical, curved, water-proof, bluetooth и т.д.
5. Цена округляется ВВЕРХ до ближайших 5 сомов (422 → 425). Без наценки.
6. Если товар с таким названием уже есть — обновляем цену, иначе создаём.
7. Непонятные названия — создаём товар без атрибутов с простым описанием.
8. Stock: default warehouse, quantity=10.

Использование:
    python manage.py import_price_sheet                          # импорт из ../прайс для Родиона.xlsx
    python manage.py import_price_sheet --file=path/to/file.xlsx # другой файл
    python manage.py import_price_sheet --dry-run                # показать план без записи
"""

import re
import math
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils.text import slugify as _django_slugify

from products.models import (
    Category, Brand, Attribute, Tag, Product, ProductAttribute, Stock,
)

DEFAULT_FILE = Path(__file__).resolve().parent.parent.parent.parent.parent / 'прайс для Родиона.xlsx'

# =============================================================================
# Транслитерация (как в seed_data.py)
# =============================================================================

CYRILLIC_TO_LATIN = {
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
    'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
    'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
    'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
    'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    'А': 'a', 'Б': 'b', 'В': 'v', 'Г': 'g', 'Д': 'd', 'Е': 'e', 'Ё': 'yo',
    'Ж': 'zh', 'З': 'z', 'И': 'i', 'Й': 'y', 'К': 'k', 'Л': 'l', 'М': 'm',
    'Н': 'n', 'О': 'o', 'П': 'p', 'Р': 'r', 'С': 's', 'Т': 't', 'У': 'u',
    'Ф': 'f', 'Х': 'h', 'Ц': 'ts', 'Ч': 'ch', 'Ш': 'sh', 'Щ': 'sch',
    'Ъ': '', 'Ы': 'y', 'Ь': '', 'Э': 'e', 'Ю': 'yu', 'Я': 'ya',
    '№': 'n', '«': '', '»': '', '"': '', '\'': '', '(': '', ')': '',
}


def transliterate(text: str) -> str:
    return ''.join(CYRILLIC_TO_LATIN.get(ch, ch) for ch in text)


def slugify(value: str) -> str:
    return _django_slugify(transliterate(value))


# =============================================================================
# Цены
# =============================================================================

def round_price_up(raw_price: float | int | str) -> int:
    """Округляет цену ВВЕРХ до ближайших 5 сомов. 422 → 425, 443.0 → 445."""
    try:
        value = float(str(raw_price).replace(' ', '').replace(',', '.'))
    except (TypeError, ValueError):
        return 0
    return int(math.ceil(value / 5) * 5)


def parse_price(value) -> float | None:
    """Извлекает число из ячейки. Возвращает None если ячейка не цена."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    # Пробуем float напрямую
    try:
        return float(text)
    except ValueError:
        pass
    # Из строки вида "1 243.67" / "1,5"
    cleaned = text.replace(' ', '').replace(',', '.')
    match = re.search(r'(\d+(?:\.\d+)?)', cleaned)
    if match:
        return float(match.group(1))
    return None


# =============================================================================
# Нормализация названия
# =============================================================================

TENDER_MARKERS = [
    r'\s*-\s*тендер\b',
    r'\s*[-–—]\s*тендер\b',
    r'\s*тендер\b',
    r'\s*-\s*б[\\/]у\b',
    r'\s*\(тендер\)',
    r'\s+\-?\s*N(EW)?\s*$',
    r'\s+NEW\s*$',
    r'\s+b\\.?у\b',
]

def normalize_title(raw: str) -> str:
    """Чистит название: убирает тендерные пометки, лишние пробелы, мусор."""
    title = str(raw).strip()
    for marker in TENDER_MARKERS:
        title = re.sub(marker, '', title, flags=re.IGNORECASE)
    # Заменяем символы вопроса/слэшей-разделителей
    title = title.replace('?', ',')
    # Схлопываем пробелы
    title = re.sub(r'\s+', ' ', title).strip(' -–—|')
    return title


# =============================================================================
# Бренды
# =============================================================================

# Каноническое имя → (slug, страна, сайт, [синонимы])
BRAND_CATALOG = {
    # --- Существующие в БД ---
    "Acer": ("acer", "Taiwan", "https://www.acer.com"),
    "AMD": ("amd", "USA", "https://www.amd.com"),
    "Apple": ("apple", "USA", "https://www.apple.com"),
    "Arctic": ("arctic", "Switzerland", "https://www.arctic.de"),
    "ASRock": ("asrock", "Taiwan", "https://www.asrock.com"),
    "ASUS": ("asus", "Taiwan", "https://www.asus.com"),
    "BenQ": ("benq", "Taiwan", "https://www.benq.com"),
    "be quiet!": ("be-quiet", "Germany", "https://www.bequiet.com"),
    "Cooler Master": ("cooler-master", "China", "https://www.coolermaster.com"),
    "Corsair": ("corsair", "USA", "https://www.corsair.com"),
    "Crucial": ("crucial", "USA", "https://www.crucial.com"),
    "DeepCool": ("deepcool", "China", "https://www.deepcool.com"),
    "Dell": ("dell", "USA", "https://www.dell.com"),
    "Ducky": ("ducky", "Taiwan", "https://www.duckychannel.com"),
    "EVGA": ("evga", "USA", "https://www.evga.com"),
    "Fractal Design": ("fractal-design", "Sweden", "https://www.fractal-design.com"),
    "Gigabyte": ("gigabyte", "Taiwan", "https://www.gigabyte.com"),
    "G.Skill": ("gskill", "Taiwan", "https://www.gskill.com"),
    "HP": ("hp", "USA", "https://www.hp.com"),
    "HyperX": ("hyperx", "USA", "https://www.hyperx.com"),
    "Intel": ("intel", "USA", "https://www.intel.com"),
    "Keychron": ("keychron", "China", "https://www.keychron.com"),
    "Kingston": ("kingston", "USA", "https://www.kingston.com"),
    "Lenovo": ("lenovo", "China", "https://www.lenovo.com"),
    "LG": ("lg", "South Korea", "https://www.lg.com"),
    "Lian Li": ("lian-li", "Taiwan", "https://www.lian-li.com"),
    "Logitech": ("logitech", "Switzerland", "https://www.logitech.com"),
    "MSI": ("msi", "Taiwan", "https://www.msi.com"),
    "Noctua": ("noctua", "Austria", "https://noctua.at"),
    "NVIDIA": ("nvidia", "USA", "https://www.nvidia.com"),
    "NZXT": ("nzxt", "USA", "https://www.nzxt.com"),
    "Patriot": ("patriot", "USA", "https://www.patriotmemory.com"),
    "Razer": ("razer", "USA", "https://www.razer.com"),
    "Samsung": ("samsung", "South Korea", "https://www.samsung.com"),
    "Seagate": ("seagate", "USA", "https://www.seagate.com"),
    "Seasonic": ("seasonic", "Taiwan", "https://www.seasonic.com"),
    "SteelSeries": ("steelseries", "Denmark", "https://www.steelseries.com"),
    "TeamGroup": ("teamgroup", "Taiwan", "https://www.teamgroup.com"),
    "Thermalright": ("thermalright", "China", "https://www.thermalright.com"),
    "Toshiba": ("toshiba", "Japan", "https://www.toshiba.com"),
    "Western Digital": ("western-digital", "USA", "https://www.westerndigital.com"),
    "Zowie": ("zowie", "Taiwan", "https://www.zowie.com"),
    # --- Новые бренды из прайса ---
    "A4TECH": ("a4tech", "China", "https://www.a4tech.com", ["A4Tech", "A4TECH"]),
    "ADATA": ("adata", "Taiwan", "https://www.adata.com"),
    "TP-Link": ("tp-link", "China", "https://www.tp-link.com", ["TP-LINK", "TP-Link"]),
    "HIKVISION": ("hikvision", "China", "https://www.hikvision.com", ["HIKVISION", "Hikvision"]),
    "SVEN": ("sven", "Russia", "https://www.sven.fi"),
    "Genius": ("genius", "Taiwan", "https://www.geniusnet.com", ["KYE"]),
    "Microlab": ("microlab", "China", "https://www.microlab-global.com"),
    "UGREEN": ("ugreen", "China", "https://www.ugreen.com"),
    "Zalman": ("zalman", "South Korea", "https://www.zalman.com"),
    "Defender": ("defender", "Russia", "https://defender.ru"),
    "Edifier": ("edifier", "China", "https://www.edifier.com"),
    "MikroTik": ("mikrotik", "Latvia", "https://mikrotik.com"),
    "CUDY": ("cudy", "China", "https://www.cudy.com"),
    "Tenda": ("tenda", "China", "https://www.tenda.com.cn"),
    "Winstar": ("winstar", "China", ""),
    "X-Game": ("x-game", "China", ""),
    "RivaCase": ("rivacase", "Ukraine", "https://rivacase.com"),
    "Promate": ("promate", "UAE", "https://www.promate.net"),
    "RITMIX": ("ritmix", "Russia", "https://ritmixrussia.ru"),
    "DELUX": ("delux", "China", "https://www.deluxworld.com", ["Delux"]),
    "Aerocool": ("aerocool", "Taiwan", "https://www.aerocool.com.tw"),
    "Aeromax": ("aeromax", "China", ""),
    "DTECH": ("dtech", "China", ""),
    "Philips": ("philips", "Netherlands", "https://www.philips.com"),
    "Brateck": ("brateck", "China", "https://www.brateck.com"),
    "MAXSUN": ("maxsun", "China", "https://www.maxsun.com.cn"),
    "TWINMOS": ("twinmos", "Taiwan", "https://www.twinmos.com", ["Twinmos"]),
    "FORESEE": ("foresee", "China", "https://www.foresee.com.cn", ["FORESEE"]),
    "ZOTAC": ("zotac", "China", "https://www.zotac.com"),
    "Dahua": ("dahua", "China", "https://www.dahuasecurity.com", ["DAHUA"]),
    "Epson": ("epson", "Japan", "https://www.epson.com"),
    "Canon": ("canon", "Japan", "https://www.canon.com"),
    "Jump": ("jump", "China", ""),
    "SVC": ("svc", "China", ""),
    "AWP": ("awp", "China", ""),
    "ANC": ("anc", "China", ""),
    "Delta": ("delta", "Taiwan", "https://www.delta.com"),
    "Matrix": ("matrix", "China", ""),
    "Dr.Web": ("dr-web", "Russia", "https://www.drweb.ru", ["Dr.Web"]),
    "Касперский": ("kaspersky", "Russia", "https://www.kaspersky.ru", ["Kaspersky"]),
    "COMIX": ("comix", "China", "https://www.comix.com.cn"),
    "KIOXIA": ("kioxia", "Japan", "https://www.kioxia.com"),
    "TWSC": ("twsc", "China", ""),
    "SOMiC": ("somic", "China", "https://www.somic.com"),
    "UNIVIEW": ("uniview", "China", "https://www.uniview.com"),
    "AOC": ("aoc", "China", "https://www.aoc.com"),
    "Xiaomi": ("xiaomi", "China", "https://www.mi.com"),
    "Share": ("share", "China", ""),
    "Щипцы": None,  # не бренд
}

# Обратный индекс: токен → каноническое имя
BRAND_TOKEN_MAP = {}
for _canon, _meta in BRAND_CATALOG.items():
    if _meta is None:
        continue
    _slug, _country, _site = _meta[0], _meta[1], _meta[2]
    _syn = _meta[3] if len(_meta) > 3 else []
    tokens = [_canon] + list(_syn)
    for _tok in tokens:
        BRAND_TOKEN_MAP[_tok.lower()] = _canon

# Многословные бренды — ищем раньше однословных
MULTIWORD_BRANDS = sorted(
    [b for b in BRAND_TOKEN_MAP if ' ' in b or '-' in b or '/' in b],
    key=len, reverse=True,
)


def detect_brand(title: str) -> str | None:
    """Определяет бренд в названии (в начале или внутри). Возвращает каноническое имя или None."""
    lower = title.lower()
    padded = f' {lower} '

    # Сначала многословные (по всему названию)
    for token in MULTIWORD_BRANDS:
        if token in padded:
            return BRAND_TOKEN_MAP[token]

    # Затем однословные
    # Сначала проверяем первое слово
    first_word = lower.split()[0] if lower.split() else ''
    if first_word in BRAND_TOKEN_MAP:
        return BRAND_TOKEN_MAP[first_word]

    # Затем ищем однословный бренд в любом месте, но только как целое слово
    words = set(lower.split())
    for word in words:
        if word in BRAND_TOKEN_MAP:
            return BRAND_TOKEN_MAP[word]
    return None


# =============================================================================
# Секции прайса → категории
# =============================================================================

# Канонические пути категорий. Родители создаются при необходимости.
SECTION_DEFAULT_CATEGORY = ('Аксессуары', 'Кабели и переходники')

# Секция (без учёта регистра) → категория
SECTION_MAP = {
    'cable': ('Аксессуары', 'Кабели и переходники'),
    'case': ('Комплектующие', 'Корпуса'),
    'cd-drive': ('Аксессуары', 'Оптические приводы CD/DVD'),
    'cooler': ('Комплектующие', 'Охлаждение'),
    'filter': ('Аксессуары', 'Сетевые фильтры'),
    'flash cards': ('Аксессуары', 'Флеш-накопители'),
    'hdd': ('Комплектующие', 'Накопители SSD'),
    'keyboard': ('Периферия', 'Клавиатуры'),
    'mainboard': ('Комплектующие', 'Материнские платы'),
    'memory': ('Комплектующие', 'Оперативная память'),
    'modem': ('Сетевое оборудование',),
    'monitors tft': ('Мониторы',),
    'mouse and pad': ('Периферия',),
    'multimedia': ('Аксессуары',),
    'network': ('Сетевое оборудование',),
    'notebook': ('Компьютеры', 'Рабочие ноутбуки'),
    'power supply': ('Комплектующие', 'Блоки питания'),
    'printer': ('МФУ и принтеры', 'Цветные МФУ и принтеры'),
    'processor': ('Комплектующие', 'Процессоры'),
    'speakers': ('Аксессуары', 'Акустика'),
    'ups': ('Аксессуары', 'Источники бесперебойного питания'),
    'video': ('Комплектующие', 'Видеокарты'),
    'аккумуляторы для ups': ('Аксессуары', 'Источники бесперебойного питания'),
    'антивирусы и программное обеспечение': ('Программное обеспечение',),
    'кронштейны': ('Аксессуары', 'Кронштейны и подставки'),
    'мебель': ('Офисная мебель',),
    'моноблок': ('Компьютеры', 'Моноблоки'),
    'оборудование офисное': ('Аксессуары', 'Офисное оборудование'),
    'планшеты и сумки для планшетов': ('Планшеты',),
    'подставки для ноутбуков': ('Аксессуары', 'Охлаждающие подставки'),
    'разное': ('Аксессуары',),
    'сумки для ноутбуков и камер': ('Аксессуары', 'Чехлы и сумки'),
}

# Секции, где категория зависит от содержимого названия
KEYWORD_CATEGORY_OVERRIDES = {
    'mouse and pad': [
        (r'pad|коврик|microsoft', ('Периферия', 'Коврики')),
        (r'mouse|мышь|mice|g-?\d|v-track|rom-|wm-|xm-|dx-|nx-|ms-|mk-', ('Периферия', 'Мыши')),
    ],
    'multimedia': [
        (r'web cam|веб-камера|веб камера', ('Аксессуары', 'Веб-камеры')),
        (r'наушник|headset|микрофон|microphone|подставка для наушников', ('Периферия', 'Гарнитуры')),
    ],
    'network': [
        (r'cable|кабель|utp', ('Аксессуары', 'Кабели и переходники')),
        (r'connector|сплиттер|сетевой двойник|щипцы|адаптер usb', ('Аксессуары', 'Кабели и переходники')),
    ],
    'hdd': [
        (r'ssd|накопитель|nvme', ('Комплектующие', 'Накопители SSD')),
        (r'hdd|жестк|1tb|western digital', ('Комплектующие', 'Жёсткие диски')),
    ],
    'printer': [
        (r'mfu|принтер|printer|лазер', ('МФУ и принтеры', 'Цветные МФУ и принтеры')),
        (r'картридж|чернил|тонер', ('МФУ и принтеры', 'Картриджи, чернила и пр.')),
    ],
}

# Для секции flаш cards — все в флеш-накопители
# Для video — видеокарты

CATEGORY_HINTS = {
    ('Аксессуары', 'Флеш-накопители'): 'flash-storage',
    ('Аксессуары', 'Сетевые фильтры'): 'power-filters',
    ('Аксессуары', 'Акустика'): 'audio-speakers',
    ('Аксессуары', 'Источники бесперебойного питания'): 'ups',
    ('Аксессуары', 'Кронштейны и подставки'): 'mounts-stands',
    ('Аксессуары', 'Охлаждающие подставки'): 'cooling-stands',
    ('Аксессуары', 'Офисное оборудование'): 'office-equipment',
    ('Аксессуары', 'Оптические приводы CD/DVD'): 'optical-drives',
    ('Сетевое оборудование',): 'networking',
    ('Офисная мебель',): 'office-furniture',
    ('Программное обеспечение',): 'software',
    ('Планшеты',): 'tablets',
    ('Компьютеры', 'Моноблоки'): 'all-in-one-pcs',
    ('Комплектующие', 'Жёсткие диски'): 'hdd',
}

# Слаг обычной категории = slugify последнего имени, но для существующих используем известные slug-и
KNOWN_SLUGS = {
    'Комплектующие': 'components', 'Процессоры': 'processors', 'Видеокарты': 'videocards',
    'Материнские платы': 'motherboards', 'Оперативная память': 'ram',
    'Накопители SSD': 'ssd', 'Жёсткие диски': 'hdd', 'Блоки питания': 'psu',
    'Корпуса': 'cases', 'Охлаждение': 'cooling', 'Компьютеры': 'computers',
    'Игровые ПК': 'gaming-pcs', 'Офисные ПК': 'office-pcs', 'Рабочие станции': 'workstations',
    'Игровые ноутбуки': 'gaming-laptops', 'Ультрабуки': 'ultrabooks', 'Рабочие ноутбуки': 'work-laptops',
    'Мониторы': 'monitors', 'Игровые мониторы': 'gaming-monitors',
    'Профессиональные мониторы': 'professional-monitors', 'Офисные мониторы': 'office-monitors',
    'Периферия': 'peripherals', 'Клавиатуры': 'keyboards', 'Мыши': 'mice',
    'Гарнитуры': 'headsets', 'Коврики': 'mousepads', 'МФУ и принтеры': 'mfu-printers',
    'Цветные МФУ и принтеры': 'color-mfu-printers',
    'Картриджи, чернила и пр.': 'cartridges-inks', 'Аксессуары': 'accessories',
    'Кабели и переходники': 'cables-adapters', 'Чехлы и сумки': 'cases-bags',
    'Веб-камеры': 'webcams',
}


def category_slug(name: str) -> str:
    """Возвращает известный slug либо генерирует."""
    return KNOWN_SLUGS.get(name, slugify(name))


# =============================================================================
# Извлечение характеристик (EAV) и тегов из названия
# =============================================================================

SCREEN_RE = re.compile(r'(\d+(?:[.,]\d+)?)["”]')
RESOLUTION_RE = re.compile(r'(\d{3,4})[xх×](\d{3,4})')
REFRESH_RE = re.compile(r'(\d{2,3})\s*[Hh][Zz]')
PANEL_RE = re.compile(r'\b(IPS|VA|OLED|TN|Nano IPS|Fast IPS)\b', re.IGNORECASE)
CPU_RE = re.compile(
    r'\b(i[3-9]-\d{4,5}[A-Z0-9]*|Ryzen\s?\d(?:\s?\d{4}[A-Z0-9]*)?|'
    r'N\d{4}|Athlon\s?Silver\s?\d{4}[A-Z0-9]*|Celeron|Pentium|'
    r'Core\s?i[3-9]-\d{4,5}[A-Z0-9]*)\b',
    re.IGNORECASE,
)
RAM_RE = re.compile(r'(\d{2,3})\s*GB(?:\s*(?:DDR[345]|LPDDR[345]|SODIMM))?', re.IGNORECASE)
SSD_RE = re.compile(r'(\d{1,3})\s*GB\s*SSD|SSD(?:\s*\w+)?\s*(\d{1,4})\s*GB|(\d{1,2})TB\s*SSD', re.IGNORECASE)
VRAM_RE = re.compile(r'(\d{1,2})\s*GB\s*GDDR|GDDR\d(?:\s*X)?\s*(\d{1,2})\s*GB', re.IGNORECASE)
SOCKET_RE = re.compile(r'\b(LGA\d{3,4}|AM[45])\b', re.IGNORECASE)
DDR_RE = re.compile(r'\b(DDR[345])\b', re.IGNORECASE)
PSU_WATT_RE = re.compile(r'(\d{3,4})\s*[Ww]')
STORAGE_TYPE_RE = re.compile(r'\b(SSD|HDD|NVMe|M\.2 SATA)\b', re.IGNORECASE)
FORM_FACTOR_RE = re.compile(r'\b(M\.2\s?2280|M\.2\s?2230|2\.5["”]|3\.5["”]|mATX|ATX|Mini-ITX)\b', re.IGNORECASE)
USB_VER_RE = re.compile(r'USB\s*[23](?:\.\d)?', re.IGNORECASE)

BOOLEAN_FEATURES = [
    (r'подсветк|rgb|\bled\b|neon', 'backlight'),
    (r'беспроводн|wireless|bluetooth|\bbt\b|wi-fi|wifi', 'wireless'),
    (r'mechanical|механик|mecha-like', 'mechanical'),
    (r'curved|изогнут', 'curved'),
]

TAG_PATTERNS = [
    (r'gaming|игров|game', 'gaming'),
    (r'wireless|беспроводн', 'wireless'),
    (r'rgb|подсветк|neon', 'rgb'),
    (r'mechanical|механик|mecha-like', 'mechanical'),
    (r'curved|изогнут', 'curved'),
    (r'water-proof|водонепрониц', 'water-proof'),
    (r'bluetooth|\bbt\b', 'bluetooth'),
    (r'wi-fi|wifi', 'wifi'),
    (r'gigabit|ethernet', 'gigabit'),
    (r'mesh', 'mesh'),
    (r'usb-c|type-c', 'usb-c'),
    (r'4k|uhd', '4k'),
    (r'fhd|full hd|1920', 'fhd'),
    (r'2k|qhd|2560|1440', 'qhd'),
    (r'ips', 'ips'),
    (r'va\b', 'va'),
    (r'oled', 'oled'),
    (r'nvme|pcie', 'nvme'),
    (r'tender|тендер', 'tender'),
    (r'беспроводной|беспроводная', 'wireless'),
]


def extract_attributes(title: str, section_key: str) -> dict:
    """
    Возвращает словарь {имя_атрибута: (тип, значение)}.
    Правила извлечения из текстового названия.
    """
    attrs: dict = {}

    # --- Диагональ экрана ---
    m = SCREEN_RE.search(title)
    if m:
        value = m.group(1).replace(',', '.')
        if section_key in ('monitors tft', 'notebook', 'моноблок', 'планшеты и сумки для планшетов'):
            if section_key in ('monitors tft',):
                attrs['Диагональ экрана'] = ('str', value)
            else:
                attrs['Диагональ ноутбука'] = ('str', value)

    # --- Разрешение ---
    m = RESOLUTION_RE.search(title)
    if m:
        attrs['Разрешение'] = ('str', f"{m.group(1)}x{m.group(2)}")

    # --- Частота обновления ---
    m = REFRESH_RE.search(title)
    if m and section_key in ('monitors tft',):
        attrs['Частота обновления'] = ('int', int(m.group(1)))

    # --- Тип матрицы ---
    m = PANEL_RE.search(title)
    if m:
        panel = 'OLED' if m.group(1).upper() == 'OLED' else m.group(1).upper()
        attrs['Тип матрицы'] = ('enum', panel)

    # --- Изогнутый ---
    if re.search(r'curved|изогнут', title, re.IGNORECASE):
        attrs['Изогнутый'] = ('bool', True)

    # --- CPU ---
    m = CPU_RE.search(title)
    if m:
        attrs['Процессор (модель)'] = ('str', m.group(1).upper())

    # --- VRAM (видеокарты) ---
    m = VRAM_RE.search(title)
    if m:
        vram = int(m.group(1) or m.group(2))
        attrs['Объём видеопамяти'] = ('int', vram)

    # --- ОЗУ: секции memory/notebook, или "NGB DDR", или первое NGB у ноутбуков ---
    is_ram_context = section_key in ('memory', 'notebook', 'моноблок')
    if is_ram_context:
        ram_m = RAM_RE.search(title)
        if ram_m and 'ssd' not in title.lower()[:ram_m.end()].lower():
            # Первое "N GB" рядом с DDR или в memory/notebook контексте — RAM
            if 'DDR' in title.upper() or section_key == 'memory' or 'ssd' not in title.lower():
                attrs['Объём ОЗУ'] = ('int', int(ram_m.group(1)))

    # --- SSD ---
    m = SSD_RE.search(title) or re.search(r'(\d{1,3})\s*GB(?:,|\s|$)', title)
    if m and section_key in ('notebook', 'моноблок', 'hdd', 'ssd'):
        if 'SSD' in title.upper() or 'NVMe' in title.upper():
            val = m.group(1) or m.group(2) or m.group(3)
            if val:
                attrs['Объём SSD'] = ('int', int(val))

    # --- Сокет ---
    m = SOCKET_RE.search(title)
    if m:
        socket_val = m.group(1).upper()
        if socket_val.startswith('LGA'):
            socket_val = f'LGA{socket_val[3:]}'
        attrs['Сокет'] = ('enum', socket_val)

    # --- Тип ОЗУ ---
    m = DDR_RE.search(title)
    if m:
        attrs['Тип ОЗУ'] = ('enum', m.group(1).upper())

    # --- Мощность БП ---
    if section_key == 'power supply':
        m = PSU_WATT_RE.search(title)
        if m:
            attrs['Мощность БП'] = ('int', int(m.group(1)))

    # --- Тип накопителя ---
    m = STORAGE_TYPE_RE.search(title)
    if m and section_key in ('hdd', 'flash cards', 'notebook'):
        st = re.sub(r'\s+', ' ', m.group(1)).upper()
        attrs['Тип накопителя'] = ('enum', st)

    # --- Форм-фактор ---
    m = FORM_FACTOR_RE.search(title)
    if m:
        ff = re.sub(r'\s+', ' ', m.group(1))
        attrs['Форм-фактор'] = ('enum', ff)

    # --- Булевы ---
    for pattern, attr_name in BOOLEAN_FEATURES:
        if re.search(pattern, title, re.IGNORECASE):
            attrs.setdefault(attr_name, ('bool', True))

    return attrs


def extract_tags(title: str) -> set[str]:
    """Возвращает множество слагов тегов."""
    tags = set()
    lower = title.lower()
    for pattern, tag in TAG_PATTERNS:
        if re.search(pattern, lower):
            tags.add(tag)
    return tags


def build_description(name: str, section: str, brand: str | None) -> str:
    """Генерирует описание товара по шаблону."""
    section_desc = {
        'cable': 'кабель и переходник',
        'case': 'корпус для ПК',
        'cd-drive': 'оптический привод',
        'cooler': 'система охлаждения компьютера',
        'filter': 'сетевой фильтр',
        'flash cards': 'карта памяти или USB-флеш-накопитель',
        'hdd': 'накопитель данных',
        'keyboard': 'клавиатура',
        'mainboard': 'материнская плата',
        'memory': 'оперативная память',
        'modem': 'сетевой модем',
        'monitors tft': 'монитор',
        'mouse and pad': 'компьютерная мышь или коврик',
        'multimedia': 'мультимедийное устройство',
        'network': 'сетевое оборудование',
        'notebook': 'ноутбук',
        'power supply': 'блок питания',
        'printer': 'принтер или МФУ',
        'processor': 'процессор',
        'speakers': 'акустическая система',
        'ups': 'источник бесперебойного питания',
        'video': 'видеокарта',
        'аккумуляторы для ups': 'аккумулятор для ИБП',
        'антивирусы и программное обеспечение': 'программное обеспечение',
        'кронштейны': 'кронштейн или подставка',
        'мебель': 'офисная мебель',
        'моноблок': 'моноблок',
        'оборудование офисное': 'офисное оборудование',
        'планшеты и сумки для планшетов': 'планшет или аксессуар для планшета',
        'подставки для ноутбуков': 'подставка для ноутбука',
        'разное': 'компьютерный аксессуар',
        'сумки для ноутбуков и камер': 'сумка или рюкзак для ноутбука',
    }.get(section, 'компьютерный товар')

    if brand:
        return (
            f"Оригинальный {brand} — {section_desc}. "
            f"Качественное решение для дома и офиса. Гарантия до 2 лет, доставка по Караколу."
        )
    return (
        f"{name} — {section_desc}. "
        f"Гарантия до 2 лет, доставка по Караколу."
    )


# =============================================================================
# Команда
# =============================================================================

class Command(BaseCommand):
    help = 'Импортирует прайс-лист поставщика (xlsx: название + цена) в каталог.'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=str(DEFAULT_FILE),
                            help='Путь к xlsx-файлу прайса.')
        parser.add_argument('--dry-run', action='store_true',
                            help='Показать план обработки без записи в БД.')
        parser.add_argument('--limit', type=int, default=0,
                            help='Ограничить количество обработанных строк (для отладки).')

    # ------------------------------------------------------------------
    # Работа с категориями
    # ------------------------------------------------------------------

    def _get_or_create_category(self, path: tuple, dry_run: bool):
        """Создаёт (или возвращает существующую) категорию по пути ('Родитель', 'Дочка')."""
        parent = None
        node = None
        for name in path:
            slug = category_slug(name)
            node, _ = Category.objects.get_or_create(
                slug=slug,
                defaults={'name': name, 'parent': parent},
            )
            parent = node
        return node

    def _category_for(self, section_key: str, title: str, dry_run: bool):
        """Определяет категорию для строки прайса."""
        if section_key in KEYWORD_CATEGORY_OVERRIDES:
            for pattern, path in KEYWORD_CATEGORY_OVERRIDES[section_key]:
                if re.search(pattern, title, re.IGNORECASE):
                    return self._get_or_create_category(path, dry_run)
            # Не распознано — используем дефолт секции
        path = SECTION_MAP.get(section_key, SECTION_DEFAULT_CATEGORY)
        return self._get_or_create_category(path, dry_run)

    # ------------------------------------------------------------------
    # Работа с брендами
    # ------------------------------------------------------------------

    def _get_or_create_brand(self, brand_name: str, dry_run: bool):
        meta = BRAND_CATALOG.get(brand_name)
        if meta is None:
            meta = ('generic', '', '')
        slug = meta[0] if meta else slugify(brand_name)
        country = meta[1] if meta and len(meta) > 1 else ''
        website = meta[2] if meta and len(meta) > 2 else ''
        brand, _ = Brand.objects.get_or_create(
            slug=slug,
            defaults={'name': brand_name, 'country': country, 'website': website},
        )
        return brand

    # ------------------------------------------------------------------
    # Характеристики (EAV)
    # ------------------------------------------------------------------

    def _get_or_create_attribute(self, name: str, attr_type: str, unit: str = ''):
        """Ищет атрибут сначала по имени (без учёта регистра), затем по слагу."""
        # Сначала точное имя — в БД могут быть атрибуты с тем же именем,
        # но слаг из транслитерации может не совпадать (например "form-factor" vs "form-faktor")
        attribute = Attribute.objects.filter(name__iexact=name).first()
        if attribute is None:
            attr_slug = slugify(name)
            attribute = Attribute.objects.filter(slug=attr_slug).first()
        if attribute is None:
            attribute = Attribute.objects.create(
                name=name,
                slug=slugify(name),
                type=attr_type,
                unit=unit,
            )
            return attribute
        # Обновляем метаданные, если изменились
        if attribute.type != attr_type:
            attribute.type = attr_type
            attribute.save()
        if attribute.unit != unit:
            attribute.unit = unit
            attribute.save()
        return attribute

    def _set_attributes(self, product, attrs: dict):
        """Сохраняет атрибуты товара через set_typed_value."""
        for attr_name, (attr_type, value) in attrs.items():
            if value is None:
                continue
            attribute = self._get_or_create_attribute(attr_name, attr_type)
            pa, created = ProductAttribute.objects.get_or_create(
                product=product,
                attribute=attribute,
            )
            try:
                pa.set_typed_value(value)
                pa.save()
            except (TypeError, ValueError):
                continue

    # ------------------------------------------------------------------
    # Товары
    # ------------------------------------------------------------------

    def _normalize_sku_slug(self, title: str, existing_slugs: set) -> str:
        base = slugify(title)
        candidate = base
        i = 1
        while candidate in existing_slugs or Product.objects.filter(slug=candidate).exists():
            i += 1
            candidate = f'{base}-{i}'
        return candidate

    def _upsert_product(self, title: str, price: int, section_key: str,
                        brand_name: str | None, attrs: dict, tags: set, dry_run: bool):
        """Создаёт или обновляет товар. Возвращает (status, product)."""
        cat = self._category_for(section_key, title, dry_run)
        brand = self._get_or_create_brand(brand_name, dry_run) if brand_name else None

        # Ищем существующий товар по нормализованному названию
        existing = Product.objects.filter(name__iexact=title).first()

        if existing and not dry_run:
            changed = False
            if existing.price != price:
                existing.price = price
                changed = True
            # Дозаполняем недостающую информацию у существующих товаров
            if brand and not existing.brands.filter(pk=brand.pk).exists():
                existing.brands.add(brand)
            if not existing.categories.filter(pk=cat.pk).exists():
                existing.categories.add(cat)
            tags_to_add = []
            for tag_slug in tags:
                if not existing.feature_tags.filter(slug=tag_slug).exists():
                    tag, _ = Tag.objects.get_or_create(slug=tag_slug, defaults={'name': tag_slug})
                    tags_to_add.append(tag)
            if tags_to_add:
                existing.feature_tags.add(*tags_to_add)
            self._set_attributes(existing, attrs)
            if changed:
                existing.save()
            Stock.objects.get_or_create(
                product=existing,
                warehouse='default',
                defaults={'quantity': 10, 'reserved': 0},
            )
            return ('updated', existing)

        if existing and dry_run:
            return ('would_update', existing)

        # Создаём
        if dry_run:
            return ('would_create', None)

        product = Product.objects.create(
            name=title,
            slug=self._normalize_sku_slug(title, set()),
            description=build_description(title, section_key, brand_name),
            price=price,
        )
        product.categories.add(cat)
        if brand:
            product.brands.add(brand)
        for tag_slug in tags:
            tag, _ = Tag.objects.get_or_create(slug=tag_slug, defaults={'name': tag_slug})
            product.feature_tags.add(tag)
        self._set_attributes(product, attrs)
        Stock.objects.get_or_create(
            product=product,
            warehouse='default',
            defaults={'quantity': 10, 'reserved': 0},
        )
        return ('created', product)

    def _process_rows(self, rows, current_section, stats, brand_stats, section_stats, processed, limit, dry_run):
        """Обрабатывает строки прайса (внутри транзакции)."""
        for name, price in rows:
            normalized_name = name.strip()

            # Секция-заголовок (без цены)
            if price is None:
                key = normalized_name.lower()
                if key in SECTION_MAP:
                    current_section = key
                    stats['skipped_header'] += 1
                    self.stdout.write(f'\n📁 Секция: {normalized_name}')
                    continue
                # Первая группа до первого заголовка — кабели
                current_section = 'cable'
                continue

            # У нас есть цена — товар
            title = normalize_title(normalized_name)
            if not title:
                stats['no_price'] += 1
                continue

            rounded = round_price_up(price)
            if rounded <= 0:
                stats['zero_price'] += 1
                continue

            brand_name = detect_brand(title)
            attrs = extract_attributes(title, current_section)
            tags = extract_tags(title)

            if brand_name:
                brand_stats[brand_name] = brand_stats.get(brand_name, 0) + 1
            else:
                brand_stats['(без бренда)'] = brand_stats.get('(без бренда)', 0) + 1
            section_stats[current_section] = section_stats.get(current_section, 0) + 1

            status, product = self._upsert_product(
                title, rounded, current_section, brand_name, attrs, tags, dry_run,
            )
            if status in stats:
                stats[status] += 1

            prefix = {
                'created': '✓',
                'updated': '↻',
                'would_create': '○',
                'would_update': '⊙',
            }.get(status, '·')
            self.stdout.write(
                f'   {prefix} [{title[:70]}] → {rounded} сом'
                f'{f" | {brand_name}" if brand_name else ""}'
                f' | атрибутов: {len(attrs)} | тегов: {len(tags)}'
            )

            processed += 1
            if limit > 0 and processed >= limit:
                self.stdout.write(self.style.WARNING(f'   ⏹ Лимит {limit} достигнут.'))
                break

    # ------------------------------------------------------------------
    # Чтение xlsx
    # ------------------------------------------------------------------

    def _read_sheet_rows(self, file_path: Path):
        """Читает xlsx и возвращает список (name, price|None)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.stderr.write(self.style.ERROR(
                'openpyxl не установлен. Выполни: pip install openpyxl'
            ))
            return []

        wb = load_workbook(str(file_path), data_only=True, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            name = row[0]
            price_raw = row[1] if len(row) > 1 else None
            if name is None or str(name).strip() == '':
                continue
            price = parse_price(price_raw)
            rows.append((str(name).strip(), price))
        wb.close()
        return rows

    # ------------------------------------------------------------------
    # Обработка
    # ------------------------------------------------------------------

    def handle(self, *args, **options):
        file_path = Path(options['file'])
        dry_run = options['dry_run']
        limit = options['limit']

        if not file_path.exists():
            self.stderr.write(self.style.ERROR(f'Файл не найден: {file_path}'))
            return

        self.stdout.write(f'📄 Импорт прайса: {file_path}')
        if dry_run:
            self.stdout.write(self.style.WARNING('🧪 РЕЖИМ DRY-RUN — запись в БД не выполняется'))

        rows = self._read_sheet_rows(file_path)
        if not rows:
            self.stderr.write(self.style.ERROR('Файл пуст или не удалось прочитать.'))
            return
        self.stdout.write(f'   Найдено строк: {len(rows)}')

        current_section = 'default'
        stats = {
            'created': 0, 'updated': 0, 'skipped_header': 0,
            'would_create': 0, 'would_update': 0,
            'no_price': 0, 'zero_price': 0,
        }
        brand_stats: dict = {}
        section_stats: dict = {}
        processed = 0

        # Весь импорт в одной транзакции — при ошибке откатывается всё
        from django.db import transaction
        try:
            with transaction.atomic():
                self._process_rows(rows, current_section, stats, brand_stats,
                                   section_stats, processed, limit, dry_run)
        except Exception as e:
            self.stderr.write(self.style.ERROR(f'❌ Ошибка импорта (транзакция откачена): {e}'))
            raise

        # Итог
        self.stdout.write('\n' + '=' * 60)
        if dry_run:
            self.stdout.write(self.style.SUCCESS('ИТОГ (dry-run):'))
            self.stdout.write(f'   Будет создано: {stats["would_create"]}')
            self.stdout.write(f'   Будет обновлено: {stats["would_update"]}')
        else:
            self.stdout.write(self.style.SUCCESS('ИТОГ:'))
            self.stdout.write(f'   Создано товаров: {stats["created"]}')
            self.stdout.write(f'   Обновлено цен: {stats["updated"]}')
        self.stdout.write(f'   Секций-заголовков: {stats["skipped_header"]}')
        self.stdout.write(f'   Строк без цены (не товары): {stats["no_price"]}')
        self.stdout.write(f'   Строк с нулевой ценой: {stats["zero_price"]}')

        if section_stats:
            self.stdout.write('\n📊 По секциям:')
            for sec, cnt in sorted(section_stats.items(), key=lambda x: -x[1]):
                self.stdout.write(f'   {sec}: {cnt}')

        if brand_stats:
            self.stdout.write('\n🏷️ По брендам (топ-15):')
            for br, cnt in sorted(brand_stats.items(), key=lambda x: -x[1])[:15]:
                self.stdout.write(f'   {br}: {cnt}')
